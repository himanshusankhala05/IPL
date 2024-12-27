from flask import Flask, jsonify, render_template, request, redirect, send_from_directory, url_for, session
import openpyxl
import pandas as pd
import os



app = Flask(__name__)
app.secret_key = 'your_secret_key'  # For session management


# File paths
CURRENT_STATE_FILE = 'current_state.txt'
CONTESTANT_PURSE_FILE = 'contestants_purse_value.txt'
USERS_LIST_FILE = 'users_list.xlsx'
AUCTION_DATA_FILE = 'auction_data.xlsx'




# Initialize current state
def initialize_state():
    if not os.path.exists(CURRENT_STATE_FILE):
        with open(CURRENT_STATE_FILE, 'w') as file:
            file.write('current_player_index:0\ncurrent_bid:0\nnew_bid:0\ncurrent_bidder:None\n')

# Load current state
def load_state():
    state = {}
    with open(CURRENT_STATE_FILE, 'r') as file:
        for line in file:
            key, value = line.strip().split(':', 1)
            state[key] = int(value) if value.isdigit() else value
    return state

# Update current state
def update_state(state):
    with open(CURRENT_STATE_FILE, 'w') as file:
        for key, value in state.items():
            file.write(f"{key}:{value}\n")



# Load player data from Excel file
def load_players():
    wb = openpyxl.load_workbook('auction_data.xlsx')
    sheet = wb.active
    players = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        player = {
            'sr_no': row[0],
            'name': row[1],
            'base_price': row[2],
            'type': row[4],
            'overseas': row[5],
            'wicketkeeper': row[6],
            'capped_uncapped': row[3],
        }
        players.append(player)
    return players

# Save the auction record to Excel file
def save_auction_record(contestant, player, bid_amount):
    wb = openpyxl.load_workbook('users_list.xlsx')
    if contestant not in wb.sheetnames:
        wb.create_sheet(contestant)
    sheet = wb[contestant]
    new_row = [
        player['name'], player['type'], player['overseas'], player['wicketkeeper'],
        player['capped_uncapped'], bid_amount
    ]
    sheet.append(new_row)
    wb.save('users_list.xlsx')

def load_contestants_purse():
    contestants_purse = {}
    with open("contestants_purse_value.txt", "r") as file:
        for line in file:
            # Split each line into key and value
            if ':' in line:
                key, value = line.split(':', 1)
                contestants_purse[key.strip()] = int(value.strip())
    return contestants_purse


def update_contestants_purse(contestants_purse):
    with open("contestants_purse_value.txt", "w") as file:
        for key, value in contestants_purse.items():
            file.write(f"{key}: {value}\n")



# Initialize contestants' purse
#contestants_purse = {
#    'Abhishekh S': 1000000000,  # Example: Contestant's purse size (in INR)
#    'Himanshu S': 1000000000,
#    'Onkar L': 1000000000,  # Example: Contestant's purse size (in INR)
#    'Skip':100000000000000,
#}



@app.route('/')
def index():
    initialize_state()
    state = load_state()
    print(state)
    # Load players and get current player index from session
    players_data = load_players()
    #current_player_index = session.get('current_player_index', 0)
    current_player_index = state['current_player_index']
    # Check if there are any players left
    if current_player_index >= len(players_data):
        return "Auction is over!"
    
    # Get the current player to display
    current_player = players_data[current_player_index]
    players_remaining = len(players_data) - current_player_index
    # Initialize current bid (starting with base price)
    #current_bid = session.get('current_bid', 0)
    current_bid=state['current_bid']
    #new_bid = session.get('new_bid', current_player['base_price']) 
    if state['new_bid']==0:
        state['new_bid'] = current_player['base_price']
        update_state(state)
    new_bid = state['new_bid']
    #current_bidder = session.get('current_bidder', None)
    current_bidder = state['current_bidder']

    contestants_purse = load_contestants_purse()
    
    return render_template('index.html', player=current_player, purse_sizes=contestants_purse, 
                           current_bid=current_bid, current_bidder=current_bidder, new_bid=new_bid, current_player_index = current_player_index + 1, players_remaining=players_remaining)

@app.route('/players')
def players():
    #current_player_index = session.get('current_player_index', 0)
    state = load_state()
    current_player_index = state['current_player_index']
    data = None
    
    
    df = pd.read_excel('auction_data.xlsx')
    df = df.iloc[current_player_index:]
    data = df.to_dict(orient="records")

    #current_player_index = session.get('current_player_index', 0)

    players_remaining = len(data)
    
    columns = df.columns.tolist()
    return render_template("players.html", data=data, columns=columns, players_remaining=players_remaining)


@app.route('/contestants', methods=["GET", "POST"])
def contestants():
    sheets = []
    data = None
    columns = []
    selected_sheet = None

    
    
    excel_file = pd.ExcelFile("users_list.xlsx")
    sheets = excel_file.sheet_names

    # Check if a sheet is selected
    selected_sheet = request.form.get("sheet_name")
    if selected_sheet:
        # Read the selected sheet into a DataFrame
        df = pd.read_excel("users_list.xlsx", sheet_name=selected_sheet)

        # Convert DataFrame to a list of dictionaries for rendering
        data = df.to_dict(orient="records")
        columns = df.columns.tolist()
    

    return render_template("contestants.html", sheets=sheets, data=data, columns=columns, selected_sheet=selected_sheet)


@app.route('/place_bid', methods=['POST'])
def place_bid():
    contestant_name = request.form['contestant']
    player_sr_no = int(request.form['player_sr_no'])

    state = load_state()
    

    # Load player data
    players_data = load_players()
    selected_player = next(player for player in players_data if player['sr_no'] == player_sr_no)
    
    

    # Calculate the new bid (increase by ₹5,00,000)
    #new_bid = session.get('new_bid', selected_player['base_price'])
    new_bid = state['new_bid']
    current_bid = new_bid
    

    increment = 0
    if new_bid >= 3000000 and new_bid < 5000000:
        increment = 2000000
    elif new_bid >= 5000000 and new_bid < 20000000:
        increment = 2500000
    else:
        increment = 5000000
        


    #session['new_bid'] = session.get('new_bid', selected_player['base_price']) + increment
    #session['current_bid'] = current_bid
    #session['current_bidder'] = contestant_name

    state['new_bid'] = new_bid + increment
    state['current_bid'] = current_bid
    state['current_bidder'] = contestant_name
    update_state(state)


    contestants_purse = load_contestants_purse()
    # Update contestant's purse
    if contestants_purse[contestant_name] >= current_bid:
        #contestants[contestant_name] -= current_bid
        
        return redirect(url_for('index'))
    else:
        
        
        # Handle insufficient funds case
        return "Insufficient funds! Please try again."


@app.route('/finalize_bid', methods=['POST'])
def finalize_bid():
    # Get current bid and bidder
    
    #current_bid = session.get('current_bid', 0)
    #current_bidder = session.get('current_bidder', None)
    
    state = load_state()
    current_bid = state['current_bid']
    current_bidder = state['current_bidder'] 

    if current_bidder is None:
        return "No bids placed!"

    # Load the current player
    players_data = load_players()
    #current_player_index = session.get('current_player_index', 0)
    current_player_index = state['current_player_index']

    selected_player = players_data[current_player_index]

    # Save the auction record to Excel
    save_auction_record(current_bidder, selected_player, current_bid)

    # Move to next player
    #current_player_index += 1
    #session['current_player_index'] = current_player_index
    #session['current_bid'] = 0
    #current_player = players_data[current_player_index]
    #session['new_bid'] = current_player['base_price']
    #session['current_bidder'] = None

    state['current_player_index'] += 1
    current_player = players_data[current_player_index+1]
    state['current_bid'] = 0
    state['new_bid'] = current_player['base_price']
    state['current_bidder'] = 'None'
    update_state(state)


    contestants_purse = load_contestants_purse()

    if contestants_purse[current_bidder] >= current_bid:
        contestants_purse[current_bidder] -= current_bid
        update_contestants_purse(contestants_purse)
        
        return redirect(url_for('index'))
   
    return redirect(url_for('index'))




if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8081, debug=True)
